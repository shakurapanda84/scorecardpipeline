# data/common/analysis.py
import pandas as pd
import numpy as np
from typing import Dict, Union
import matplotlib.pyplot as plt
import seaborn as sns
import io
from openpyxl import ExcelWriter
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.utils.dataframe import dataframe_to_rows

class Analysis:
    def __init__(self, data: pd.DataFrame):
        """
        Initialize with a DataFrame containing account performance data
        
        Parameters:
        -----------
        data : pd.DataFrame
            Must contain:
            - account_id: Unique account identifier
            - observation_date: Date of observation
            - delinquency_status: Days delinquent (0=current, 30=30 days late, etc.)
        """
        required_columns = {'account_id', 'observation_date', 'delinquency_status'}
        if not required_columns.issubset(data.columns):
            missing = required_columns - set(data.columns)
            raise ValueError(f"Missing required columns: {missing}")
            
        self.data = data.sort_values(['account_id', 'observation_date'])
        
    def roll_rate_analysis(self, 
                          period: str = 'M',
                          lookback_periods: int = 12) -> pd.DataFrame:
        """
        Perform roll rate analysis across specified time periods
        
        Parameters:
        -----------
        period : str
            Time period for analysis ('M' for monthly, 'Q' for quarterly)
        lookback_periods : int
            Number of periods to analyze
            
        Returns:
        --------
        pd.DataFrame
            Roll rate matrix with transition probabilities
        """
        # Create period reference
        self.data['period'] = self.data['observation_date'].dt.to_period(period)
        
        # Get unique periods sorted
        periods = self.data['period'].dropna().sort_values().unique()
        if len(periods) < 2:
            raise ValueError("Insufficient data periods for roll rate analysis")
            
        # Create lagged status
        self.data['prev_status'] = self.data.groupby('account_id')['delinquency_status'].shift(1)
        self.data['next_status'] = self.data.groupby('account_id')['delinquency_status'].shift(-1)
        
        # Filter relevant periods
        analysis_data = self.data.dropna(subset=['prev_status', 'next_status'])
        
        # Create status buckets
        status_bins = [-1, 0, 30, 60, 90, 180, np.inf]
        analysis_data['current_bucket'] = pd.cut(analysis_data['prev_status'], bins=status_bins)
        analysis_data['next_bucket'] = pd.cut(analysis_data['next_status'], bins=status_bins)
        
        # Create roll rate matrix
        roll_matrix = pd.crosstab(
            index=analysis_data['current_bucket'],
            columns=analysis_data['next_bucket'],
            normalize='index'
        ).round(4) * 100
        
        # Calculate roll rates
        roll_matrix['total'] = roll_matrix.sum(axis=1)
        roll_matrix = roll_matrix.sort_index(ascending=False)
        
        return roll_matrix

def analyze_categorical_bads(df, column_name):
    """
    Analyze a categorical column in a DataFrame to produce grouped statistics and visualizations.

    Parameters:
    -----------
    df : pd.DataFrame
        The input DataFrame containing the data.
    column_name : str
        The name of the categorical column to analyze.

    Returns:
    --------
    pd.DataFrame
        A combined DataFrame containing the grouped statistics.
    """
    # Ensure the 'bad' column exists
    if 'bad' not in df.columns:
        raise ValueError("The DataFrame must contain a 'bad' column.")

    # Ensure the 'source' column exists
    if 'source' not in df.columns:
        raise ValueError("The DataFrame must contain a 'source' column.")

    # 1. Grouped count of bads by column value
    grouped_by_value = df.groupby(column_name).agg(
        bad_count=('bad', 'sum'),
        total_count=('bad', 'count')
    ).reset_index()
    grouped_by_value['bad_percentage'] = (grouped_by_value['bad_count'] / grouped_by_value['total_count']) * 100
    grouped_by_value['source'] = 'all'  # Add source column with 'all'

    # 2. Grouped count of bads by column value and source
    grouped_by_value_source = df.groupby([column_name, 'source']).agg(
        bad_count=('bad', 'sum'),
        total_count=('bad', 'count')
    ).reset_index()
    grouped_by_value_source['bad_percentage'] = (grouped_by_value_source['bad_count'] / grouped_by_value_source['total_count']) * 100

    # Combine the two DataFrames
    combined_df = pd.concat([grouped_by_value, grouped_by_value_source], ignore_index=True)

    # 3. Plot bar charts for each group
    for source in ['all', 'booked', 'unbooked']:
        plot_data = combined_df[combined_df['source'] == source]
        fig, ax1 = plt.subplots(figsize=(12, 6))

        # Plot bad count
        sns.barplot(x=column_name, y='bad_count', data=plot_data, ax=ax1, color='b', ci=None)
        ax1.set_title(f'Grouped Bad Counts and Mean Bad Rate by {column_name} ({source})')
        ax1.set_xlabel(column_name)
        ax1.set_ylabel('Bad Count', color='b')
        ax1.tick_params(axis='y', labelcolor='b')

        # Create a second y-axis for the mean bad rate
        ax2 = ax1.twinx()
        sns.lineplot(x=column_name, y='bad_percentage', data=plot_data, ax=ax2, color='r', marker='o')
        ax2.set_ylabel('Bad Percentage', color='r')
        ax2.tick_params(axis='y', labelcolor='r')

        plt.xticks(rotation=45)
        plt.tight_layout()
        plt.show()

    return combined_df

def analyze_and_export_to_excel(df, columns, excel_file_path):
    """
    Analyze a list of categorical columns in a DataFrame and export results to an Excel file.

    Parameters:
    -----------
    df : pd.DataFrame
        The input DataFrame containing the data.
    columns : list of str
        The list of categorical column names to analyze.
    excel_file_path : str
        The file path for the output Excel file.

    Returns:
    --------
    None
    """
    # Create a new workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove the default sheet

    for column_name in columns:
        # Perform analysis
        combined_df = analyze_categorical_bads(df, column_name)

        # Add a new sheet for each column
        ws = wb.create_sheet(title=column_name)

        # Write DataFrame to Excel
        for r_idx, row in enumerate(dataframe_to_rows(combined_df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        # Plot and save the figure to the Excel file
        row_offset = len(combined_df) + 2  # Start plotting after the data
        for source in ['all', 'booked', 'unbooked']:
            plot_data = combined_df[combined_df['source'] == source]
            fig, ax1 = plt.subplots(figsize=(12, 6))

            # Plot bad count
            sns.barplot(x=column_name, y='bad_count', data=plot_data, ax=ax1, color='b', ci=None)
            ax1.set_title(f'Grouped Bad Counts and Mean Bad Rate by {column_name} ({source})')
            ax1.set_xlabel(column_name)
            ax1.set_ylabel('Bad Count', color='b')
            ax1.tick_params(axis='y', labelcolor='b')

            # Create a second y-axis for the mean bad rate
            ax2 = ax1.twinx()
            sns.lineplot(x=column_name, y='bad_percentage', data=plot_data, ax=ax2, color='r', marker='o')
            ax2.set_ylabel('Bad Percentage', color='r')
            ax2.tick_params(axis='y', labelcolor='r')

            plt.xticks(rotation=45)
            plt.tight_layout()

            # Save the plot to a BytesIO object
            image_stream = io.BytesIO()
            plt.savefig(image_stream, format='png')
            plt.close(fig)

            # Insert the image into the Excel sheet
            image_stream.seek(0)
            img = Image(image_stream)
            img.anchor = f'G{row_offset}'  # Position the image
            ws.add_image(img)

            row_offset += 20  # Adjust the row offset for the next plot

    # Save the workbook
    wb.save(excel_file_path)
    print(f"Analysis results have been written to {excel_file_path}")

# Example usage
if __name__ == "__main__":
    # Sample data
    data = {
        'category': ['A', 'B', 'A', 'C', 'B', 'A', 'C', 'B'],
        'bad': [1, 0, 1, 0, 1, 0, 1, 0],
        'source': ['booked', 'unbooked', 'booked', 'unbooked', 'booked', 'unbooked', 'booked', 'unbooked']
    }
    df = pd.DataFrame(data)

    # List of columns to analyze
    columns_to_analyze = ['category']

    # Analyze and export to Excel
    analyze_and_export_to_excel(df, columns_to_analyze, 'analysis_results.xlsx')